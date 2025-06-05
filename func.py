# -*- coding: utf-8 -*-
"""
Created on Fri Feb 16 18:12:13 2024

@author: Jiashang
"""

# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
This grogram is used to search the potential metabolites in the providing regions of interest based on 2 data sets

Input: 
    region of interest / 5 col: upper lim, lower lim, increase/decrease, significance, FDR significance

    2 data sets: 1. all the NMR measured metabolites in a certain sample type (csf, feces, saliva, serum, sweat, urine) 
                 2. disease related metabolites list(Alzheimer's Disease, Lung Cancer, Prostate Cancer)

Output:
    the metabolites table in three different categories 

"""
import pandas as pd
import numpy as np
import os
from fractions import Fraction as frac
import xlsxwriter
import pandas.io.formats.excel
import copy
import math


def ind(array, item):
    for idx, val in np.ndenumerate(array):
        if val == item:
            return idx


# This is for removing the metabolites' ppm & region beyond defined region; ppm table0 is after the removement
def define_region(region):
    global ppm_table
    global region_table
    global ppm_table0
    global region_table0
    if ';' in region:
        region = region.split('; ')
    else:
        region = [region]
    covered_region1 = np.zeros((len(region), 2))
    for i in range(0, len(covered_region1)):
        covered_region1[i] = np.array(region[i].split('-'))
    low_region = []
    upp_region = []
    for i in range(0, covered_region1.shape[0]):
        low_region.append(min([covered_region1[i, 0], covered_region1[i, 1]]))
        upp_region.append(max([covered_region1[i, 0], covered_region1[i, 1]]))
    for m in range(0, len(ppm_table0)):
        temp_list = ppm_table0[m]
        for j in range(1, ppm_table0.shape[1]):
            out_range = 1
            if not pd.isnull(temp_list[j]):
                for h in range(0, len(low_region)):
                    if temp_list[j] >= low_region[h] and temp_list[j] <= upp_region[h]:
                        out_range = 0
                if out_range == 1:
                    temp_list[j] = float('nan')
        ppm_table0[m] = temp_list

    for m in range(0, len(region_table0)):
        temp_list = region_table0[m]
        cp_region1 = [x for x in list(region_table0[m]) if pd.isnull(x) == False][
                     1:]  # get rid of the 'nan' in the table
        for j in range(0, len(cp_region1)):
            single_region1 = cp_region1[j].replace('[', ',').replace(']', ',').split(',')[1:-1]  # get rid of the '[]'
            single_region1 = sorted(list(map(float, single_region1)))
            out_range = 1
            for h in range(0, len(low_region)):
                if low_region[h] <= single_region1[0] < upp_region[h] or low_region[h] < single_region1[1] <= \
                        upp_region[h]:
                    out_range = 0
            if out_range == 1:
                temp_list[j + 1] = float('nan')
        region_table0[m] = temp_list

    # search function, search the metabolites in the ppm region of the imp_data


def search(ppm_low_lim, ppm_upp_lim):
    global ppm_table
    output_dict = {}
    for i in range(0, len(ppm_table)):  # for each cp/row in the ppm_table
        j = 1  # start col for each ppm value w/in the cp/row
        while not pd.isnull(ppm_table[i, j]):  # while the ppm_value is an int
            if ppm_table[i, j] >= ppm_low_lim and ppm_table[i, j] <= ppm_upp_lim:
                if ppm_table[i, 0] in output_dict:  # if the cp name already in dict
                    prev_list = output_dict[ppm_table[i, 0]]
                    prev_list.append(ppm_table[i, j])
                    output_dict[ppm_table[i, 0]] = prev_list  # just add ppm value

                else:  # if cp name not in, create a new dict entry
                    output_dict[ppm_table[i, 0]] = [ppm_table[i, j]]
            j += 1
            if j >= ppm_table.shape[1]:
                break
    return output_dict


# add increase/decrease and significance information
def searchplus(low_lim, upp_lim):
    global imp_data
    output_dict_2 = {}
    prev_list = []
    for i in range(0, len(ppm_table)):
        starnumber = []
        trend = []
        FDR_list = []
        for n in range(0, imp_data.shape[0]):
            FDR = 0
            star_point = 0
            for j in range(1, ppm_table.shape[1]):
                if not pd.isnull(ppm_table[i, j]):  # while the ppm_value is an int
                    if ppm_table[i, j] >= low_lim[n] and ppm_table[i, j] <= upp_lim[n]:
                        trend.append(imp_data[n, 2])
                        if imp_data[n, 4] == '!':
                            FDR = 1
                        if imp_data[n, 3] == '*':
                            star_point = 1
                        if ppm_table[i, 0] in output_dict_2:  # if the cp name already in dict
                            prev_list = output_dict_2[ppm_table[i, 0]]
                            prev_list.append(ppm_table[i, j])
                            output_dict_2[ppm_table[i, 0]] = prev_list  # just add ppm value
                        else:  # if cp name not in, create a new dict entry
                            output_dict_2[ppm_table[i, 0]] = [ppm_table[i, j]]

                if j == ppm_table.shape[1] - 1 and not len(prev_list) == 0:
                    break
            if star_point == 1:
                starnumber.append('*')
            if FDR == 1:
                FDR_list.append('!')
        if not len(trend) == 0:
            output_dict_2[ppm_table[i, 0]] = [output_dict_2[ppm_table[i, 0]], trend, starnumber, FDR_list]

    # pick up the metabolites that may be ignored by the different trend
    # *** the trend in FDR-signifi region will override the trend in normal region***
    # *** But the trend in FDR-signifi region can not override the trend in significant region***
    # *** if the metabolites not have the FDR-signifi region, then non trend is overrided***

    for i in list(output_dict_2.keys()):
        starnumber = []
        trend = []
        FDR_list = []
        ppm_list = []
        if '!' in output_dict_2[i][3]:
            ppm_list = output_dict_2[i][0]
            for n in range(0, imp_data.shape[0]):  # the imput region
                FDR = 0
                star_point = 0
                for j in range(0, len(output_dict_2[i][0])):  # the list of the ppm
                    if output_dict_2[i][0][j] >= low_lim[n] and output_dict_2[i][0][j] <= upp_lim[n]:
                        if imp_data[n, 3] == '*':
                            star_point = 1
                        if imp_data[n, 4] == '!':
                            FDR = 1
                if star_point == 1:
                    starnumber.append('*')
                    trend.append(imp_data[n, 2])
                if FDR == 1:
                    FDR_list.append('!')
                    trend.append(imp_data[n, 2])
            output_dict_2[i] = [ppm_list, trend, starnumber, FDR_list]
    return output_dict_2


# arrange the abbreviated metabolites in alphabetical order in different categories
def forma_dict(dict0):
    global f_table2
    global abbreviation_table
    for i in list(dict0.keys()):
        me_list = dict0[i]
        me_list0 = me_list.copy()
        for j in range(0, len(me_list)):
            ab_meta = me_list0[j]
            na_meta = ab_meta
            for m in range(0, abbreviation_table.shape[0]):
                if ab_meta == abbreviation_table[m, 1]:
                    na_meta = abbreviation_table[m, 0]
            meta_index = ind(f_table2, na_meta)
            conc_type = f_table2[meta_index[0], 7]
            if conc_type == "not quantified":
                me_list.remove(ab_meta)
                me_list.append(ab_meta)
        dict0[i] = me_list
    return dict0


def sort_fractions(table):
    table = table[table[:, 0].argsort()[::-1]]
    fraction = table[0, 0]
    low_index = 0
    high_index = 0
    new_f_table = list(range(0, table.shape[1]))
    for i in range(0, len(table)):
        if table[i, 0] == fraction:
            if i == table.shape[0] - 1:
                table_to_sort = np.array(table[low_index:, :])
                table_to_sort = table_to_sort[table_to_sort[:, 1].argsort()[::-1]]
                new_f_table = np.vstack((new_f_table, table_to_sort))
            else:
                high_index = i
        else:
            high_index += 1
            table_to_sort = np.array(table[low_index:high_index, :])
            table_to_sort = table_to_sort[table_to_sort[:, 1].argsort()[::-1]]
            new_f_table = np.vstack((new_f_table, table_to_sort))
            if i == table.shape[0] - 1:
                table_to_sort = np.array([table[-1, :]])
                table_to_sort = table_to_sort[table_to_sort[:, 1].argsort()[::-1]]
                new_f_table = np.vstack((new_f_table, table_to_sort))
            low_index = i
            high_index = i
            fraction = table[i, 0]
    new_f_table = np.delete(new_f_table, 0, 0)
    new_f_table = np.delete(new_f_table, 0, 1)
    if new_f_table.shape[1] == 10:
        new_f_table = np.delete(new_f_table, 4, 1)
        new_f_table = np.delete(new_f_table, 6, 1)
    return new_f_table


def get_all_metabolites(dict1, list1):
    group_list = ['increase and significant', 'increase and not significant'
        , 'increase and FDR-signifi', 'decrease and significant'
        , 'decrease and not significant', 'decrease and FDR-signifi']
    for x in group_list:
        for y in dict1[x]:
            list1.append(y)
    list1 = list(set(list1))
    list1.sort(key=str.lower)
    return list1


def format_f_table(table):
    f_table = ["metabolites", "abbreviations", "match ratio", 'matched regions', "concentration range"]

    _outputTable = {
        "metabolites":[],
        "abbreviations":[],
        "match ratio":[],
        'matched regions':[],
        "concentration range":[]
    }

    for i in range(0, table.shape[0]):
        temp_row = []
        for j in range(0, len(metabolites_list)):
            if table[i, 2] == metabolites_list[j]:
                temp_row.append(table[i, 2])
                _outputTable['metabolites'].append(table[i, 2])

                abbre_check = 0
                for m in range(0, abbreviation_table.shape[0]):
                    if table[i, 2] == abbreviation_table[m, 0]:
                        temp_row.append(abbreviation_table[m, 1])
                        _outputTable['abbreviations'].append(abbreviation_table[m, 1])
                        abbre_check = 1

                if abbre_check == 0:
                    temp_row.append(' ')
                    _outputTable['abbreviations'].append(' ')

                temp_row.append(table[i, 1])
                _outputTable["match ratio"].append(table[i, 1])

                temp_row.append(table[i, 3])
                _outputTable["matched regions"].append(table[i, 3])

                temp_row.append(table[i, 8])
                _outputTable["concentration range"].append(table[i, 8])

                temp_row = np.array(temp_row, dtype=object)
                f_table = np.vstack((f_table, temp_row))

    return f_table,_outputTable


def abbreviate_dict(dict1):
    global abbreviation_table
    group_list = ['increase and significant', 'increase and not significant', 'increase and FDR-signifi',
                  'decrease and significant', 'decrease and not significant', 'decrease and FDR-signifi',
                  'no trend and FDR-signifi', 'no trend and not significant', 'no trend and significant']
    dict2 = {'increase and significant': [], 'increase and not significant': [], 'increase and FDR-signifi': [],
             'decrease and significant': [], 'decrease and not significant': [], 'decrease and FDR-signifi': [],
             'no trend and FDR-signifi': [], 'no trend and not significant': [], 'no trend and significant': []}
    for x in group_list:
        temp_list = []
        for i in dict1[x]:
            for m in range(0, abbreviation_table.shape[0]):
                if i == abbreviation_table[m, 0]and abbreviation_table[m,1] != ' ':
                    i = abbreviation_table[m, 1]
            temp_list.append(i)
        temp_list.sort(key=str.lower)
        dict2[x] = temp_list
    return dict2


def turn_table_to_dict(table):
    c_dict = {'increase and significant': [], 'increase and not significant': [], 'increase and FDR-signifi': [],
              'decrease and significant': [], 'decrease and not significant': [], 'decrease and FDR-signifi': [],
              'no trend and FDR-signifi': [], 'no trend and not significant': [], 'no trend and significant': []}
    for i in range(0, table.shape[0]):
        if len(table[i, 6]) != 0:
            c_dict[table[i, 6]].append(table[i, 1])
    return c_dict


def output_the_list(list11):
    global filename_categorize
    os.chdir(file_location)
    workbook = xlsxwriter.Workbook(filename_categorize)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'metabolites')
    for i in range(0, len(list11)):
        worksheet.write(i + 1, 0, list11[i])
    worksheet.set_column('A:A', 50)
    workbook.close()


# out put the result in 2nd way
def categorize_in_regions(category_dicts):  # list the metabolites with their catagory in each ppm region
    global imp_data
    global low_lim
    global upp_lim
    global abbreviation_table
    global filename_categorize
    os.chdir(file_location)
    workbook = xlsxwriter.Workbook(filename_categorize)
    worksheet = workbook.add_worksheet()
    worksheet.write(0, 0, 'region')
    worksheet.write(0, 2, 'increase')
    worksheet.write(0, 1, 'decrease')
    decrease_c = ['decrease and significant', 'decrease and not significant', 'decrease and FDR-signifi']
    increase_c = ['increase and significant', 'increase and not significant', 'increase and FDR-signifi']
    decrease_list = []
    increase_list = []
    for x in decrease_c:
        for y in category_dicts[x]:
            decrease_list.append(y)
    for x in increase_c:
        for y in category_dicts[x]:
            increase_list.append(y)
    row_number = 1
    for h in range(0, imp_data.shape[0]):
        output_dict = search(low_lim[h], upp_lim[h])
        de_temp_list = []
        in_temp_list = []
        for i in output_dict:
            for m in range(0, abbreviation_table.shape[0]):
                if i == abbreviation_table[m, 0] and abbreviation_table[m, 1] != " ":
                    i = abbreviation_table[m, 1]
            if i in decrease_list:
                de_temp_list.append(i)
            if i in increase_list:
                in_temp_list.append(i)
        worksheet.write(row_number, 0, str(upp_lim[h]) + '-' + str(low_lim[h]))
        worksheet.write(row_number, 1, ', '.join(de_temp_list))
        worksheet.write(row_number, 2, ', '.join(in_temp_list))
        row_number += 1
    workbook.close()


# output the result in 1st way
def categorize_in_groups(dict_f):  # categroized the metabolites based on the group and trend
    global filename_categorize
    global in_layer_list
    global abbreviation_table
    global f_table2

    def seqlist(list00):#sequence the metabolites by the potential concentration
        con_type_list = [">500 uM", "[5 uM, 500 uM]", "not quantified", '<5 uM', ]
        l1, l2, l3 = [[] for n in range(0, 3)]
        list01 = []
        for x in list00:
            ab_index = ind(abbreviation_table, x)
            if ab_index != None:
                ab_meta = abbreviation_table[ab_index[0], 0]
            else:
                ab_meta = x
            meta_index = ind(f_table2, ab_meta)
            conc_type = f_table2[meta_index[0], 7]
            index_c = con_type_list.index(conc_type)
            if index_c == 0:
                l1.append(x)
            if index_c == 1:
                l2.append(x)
            if index_c == 2:
                l3.append(x)
        de_list1 = [l1, l2, l3]
        for m in de_list1:
            m.sort()
            for y in m:
                list01.append(y)
        return list01

    os.chdir(file_location)
    workbook = xlsxwriter.Workbook(filename_categorize)
    worksheet = workbook.add_worksheet()
    type0 = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman'})
    type1 = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman'})
    type2 = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'font_name': 'Times New Roman'})
    worksheet.write(0, 0, 'groups')
    worksheet.write(0, 1, 'decrease')
    worksheet.write(0, 5, 'increase')
    decrease_list = []
    increase_list = []
    de_layer_list = [0]
    in_layer_list = [0]
    de_nu = 0
    in_nu = 0
    # the groups that categorized before
    de_group_list = ['decrease and FDR-signifi', 'decrease and significant',
                     'decrease and not significant']
    in_group_list = ['increase and FDR-signifi', 'increase and significant',
                     'increase and not significant']
    format_list = [type0, type1, type2]  # different formats to distinguish different categories
    row_number = 1
    layer_number = 1
    # for x in range(0,3):
    for x in range(0, 3):
        de_nu += math.ceil(len(dict_f[de_group_list[x]]) / 4)
        in_nu += math.ceil(len(dict_f[in_group_list[x]]) / 4)
        if len(dict_f[de_group_list[x]]) == 0:
            de_nu = 1
        if len(dict_f[in_group_list[x]]) == 0:
            in_nu = 1
        de_layer_list.append(de_nu)
        in_layer_list.append(in_nu)
    for x in range(0, 3):
        if de_layer_list[x + 1] < in_layer_list[x + 1]:
            gap = in_layer_list[x + 1] - de_layer_list[x + 1]
            for g in range(x + 1, len(de_layer_list)):
                de_layer_list[g] += gap
        if in_layer_list[x + 1] < de_layer_list[x + 1]:
            gap = de_layer_list[x + 1] - in_layer_list[x + 1]
            for g in range(x + 1, len(in_layer_list)):
                in_layer_list[g] += gap
    if in_layer_list[1] == in_layer_list[2]:
        in_layer_list[2] += 1
        in_layer_list[3] += 1
    de_layer_list = in_layer_list
    for x in range(0,
                   3):  # write the groups in order: FDR-signifi goes first, then significant and the last is not significant
        row_number = de_layer_list[x] + 1
        if len(dict_f[de_group_list[x]]) == 0:
            row_number += 1
        layer_number = row_number
        de_seq_list = dict_f[de_group_list[x]]
        de_seq_list = seqlist(de_seq_list)
        for i in de_seq_list:
            decrease_list.append(i)
            if len(decrease_list) == 4:  # 4 metabolites in one row
                for z in range(1, 5):
                    worksheet.write(row_number, z, decrease_list[z - 1], format_list[x])
                decrease_list = []
                row_number += 1
            if len(decrease_list) == len(dict_f[de_group_list[x]]) - 4 * (row_number - layer_number) and len(
                    decrease_list) != 0:  # row number less than 4
                for z in range(1, len(decrease_list) + 1):
                    worksheet.write(row_number, z, decrease_list[z - 1], format_list[x])
                decrease_list = []
                row_number += 1
    row_number = 1
    layer_number = 1
    for x in range(0, 3):
        row_number = in_layer_list[x] + 1
        if len(dict_f[in_group_list[x]]) == 0:
            row_number += 1
        layer_number = row_number
        in_seq_list = dict_f[in_group_list[x]]
        in_seq_list = seqlist(in_seq_list)
        for i in in_seq_list:
            increase_list.append(i)
            '''
            if row_number != 1 and in_layer_list[x] <= de_layer_list[x]:
                gap = de_layer_list[x] - row_number +1
                row_number = de_layer_list[x]+1
                in_layer_list[x+1] = in_layer_list[x+1]+gap
         '''
            if len(increase_list) == 4:  # 4 metabolites in one row
                for z in range(1, 5):
                    worksheet.write(row_number, z + 4, increase_list[z - 1], format_list[x])
                increase_list = []
                row_number += 1
            if len(increase_list) == len(dict_f[in_group_list[x]]) - 4 * (row_number - layer_number) and len(
                    increase_list) != 0:
                for z in range(1, len(increase_list) + 1):
                    worksheet.write(row_number, z + 4, increase_list[z - 1], format_list[x])
                increase_list = []
                row_number += 1
    workbook.close()


# add color for metabolites in different concentration level; ">500uM" in the front; "[5uM, 500uM]" in the middle; "not quantified" in the end
def forma_excel(filename_categorize):
    global f_table2
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    os.chdir(file_location)
    workbook = load_workbook(filename=filename_categorize)
    sheet = workbook['Sheet1']
    max_row = sheet.max_row
    max_cow = sheet.max_column
    Color = ["808080", "BFBFBF", "F2F2F2"]
    for x in range(0, 3):
        fille = PatternFill('solid', fgColor=Color[x])
        for h in range(in_layer_list[x] + 2, in_layer_list[x + 1] + 2):
            for w in range(2, 10):
                sheet.cell(row=h, column=w).fill = fille
    for i in range(2, max_row + 1):
        for j in range(2, max_cow + 1):
            if sheet.cell(i, j).value != None:
                meta_name = sheet.cell(i, j).value
                ab_index = ind(abbreviation_table, meta_name)
                if ab_index != None:
                    ab_meta = abbreviation_table[ab_index[0], 0]
                else:
                    ab_meta = meta_name
                meta_index = ind(f_table2, ab_meta)
                conc_type = f_table2[meta_index[0], 7]
                if conc_type == '<5 uM':
                    sheet.cell(i, j).value = meta_name  # + '(<5 uM)'
                if conc_type == "[5 uM, 500 uM]":
                    sheet.cell(i, j).value = meta_name  # + '([5 uM, 1 mM])'
                if conc_type == ">500 uM":
                    sheet.cell(i, j).value = meta_name  # + '(>1 mM)'
                    sheet.cell(i, j).font = Font(name='Times New Roman', color="FF0000", bold=True, italic=False)
                if conc_type == "not quantified":
                    sheet.cell(i, j).value = meta_name  # + '(not quantified)'
                    sheet.cell(i, j).font = Font(name='Times New Roman', color="0070C0", bold=False, italic=True)
    workbook.save(filename_categorize)


def UI_search():
    '''search, exports as excel file "output_cp_search.xlsx"'''
    global imp_data
    global category_dict
    global category_dict2
    global low_lim
    global upp_lim
    global f_table2
    global f_table
    global p_table
    global category_dict_p1
    global category_dict_p2
    global metabolites_list
    global category_type
    global ans
    global action
    global output_location
    global output_dict2
    '''
    identify potential metabolites from the imput_data based on their ppm peaks
    '''
    if imp_data.shape[0]>1:
        define_region(covered_region)
    low_lim = []
    upp_lim = []
    for i in range(0, imp_data.shape[0]):
        low_lim.append(min([imp_data[i, 0], imp_data[i, 1]]))
        upp_lim.append(max([imp_data[i, 0], imp_data[i, 1]]))
    output_list = []
    for n in range(0, imp_data.shape[0]):
        output_dict = search(low_lim[n], upp_lim[n])
        output_str = ""
        for j in range(0, len(output_dict)):
            output_str = str(output_str) + list(output_dict.keys())[j] + ' ' + str(
                sorted(set(list(output_dict.values())[j]))) + "\n"
        if output_str == '':
            output_str = 'none'
        output_list.append(output_str)
    '''
    read the significant and FDR data; categorize the metabolites
    '''
    output_dict2 = searchplus(low_lim, upp_lim)  # add in/decrease, significance and FDR-signifi information
    low_lim2 = np.vstack(np.array(low_lim))
    upp_lim2 = np.vstack(np.array(upp_lim))
    imp_data_ppm = np.hstack((low_lim2, upp_lim2))

    # join imp_data + array
    out_data = pd.DataFrame(imp_data_ppm, columns=['low lim', 'upp lim'])
    output_list = pd.DataFrame(output_list, columns=['compounds'])
    df = pd.concat([out_data, output_list], axis=1)
    f_table2 = ["", "", "", "", "", "", "", "", ""]
    uq_cps_dict2 = output_dict2.copy()
    category_dict = {'increase and significant': [], 'increase and not significant': [], 'increase and FDR-signifi': [],
                     'decrease and significant': [], 'decrease and not significant': [], 'decrease and FDR-signifi': [],
                     'no trend and FDR-signifi': [], 'no trend and not significant': [], 'no trend and significant': []}
    for i in range(0, len(list(uq_cps_dict2.keys()))):
        row = []
        signifi_str = 0
        trend_str = 0
        category = 0
        conc_gory = "not quantified"
        cp_name = list(uq_cps_dict2.keys())[i]
        cp_index = ind(region_table, cp_name)
        conc_index = ind(concfilter, cp_name)
        if conc_index != None:
            conc_row = conc_index[0]
            if concfilter[conc_row, 2] > 500:
                conc_gory = ">500 uM"
            if concfilter[conc_row, 2] < 500 and concfilter[conc_row, 2] > 5:
                conc_gory = "[5 uM, 500 uM]"
            if concfilter[conc_row, 2] < 5:
                conc_gory = "<5 uM"
        if cp_index != None:
            cp_row = cp_index[0]
            cp_region = [x for x in list(region_table[cp_row]) if pd.isnull(x) == False][
                        1:]  # get rid of the 'nan' in the table
            cp_region = list(set(cp_region))
            cp_region0 = [x for x in list(region_table0[cp_row]) if pd.isnull(x) == False][
                         1:]  # get rid of the 'nan' in the table
            cp_region0 = list(set(cp_region0))
            bottom_pass = 0
            if frac(len(cp_region0),
                    len(cp_region)) > 0:  # if more than [certain ratio] of the metabolites regions are removed, the metabolites will not be identified
                bottom_pass = 1
            cp_peaks = sorted(set(list(uq_cps_dict2.values())[i][0]))
            # bottom = total number of regions in region_table
            bottom = len(cp_region0) if len(cp_region0) != 0 else 1
            # top = number of ppm values that appeared in the region
            top = 0
            star_number = 0
            if bottom >= 1 and bottom_pass == 1:
                for n in range(0,
                               bottom):  # we found the ppm appears first, and according to the ppm we can calculate how many region appears
                    single_region = cp_region0[n].replace('[', ',').replace(']', ',').split(',')[
                                    1:-1]  # get rid of the '[]'
                    single_region = list(map(float, single_region))
                    count = 0
                    signifi = 0
                    for m in cp_peaks:
                        if m >= single_region[0] and m <= single_region[1]:
                            count += 1  # calculate the match ratio
                            for j in range(0, imp_data.shape[0]):
                                if m >= imp_data[j, 1] and m <= imp_data[j, 0] and imp_data[j, 3] == '*':
                                    signifi = 1
                    if signifi == 1:
                        star_number += 1
                    if count >= 1:
                        # ***At least one ppm shows up in the region, then we can say the region is picked up***
                        top += 1
            fraction = frac(top, bottom)
            str_fraction = str(top) + '/' + str(
                bottom)  # the match ratio is based on the region shows up instead of the ppm
            '''
            match ratio select standard
            '''
            if ((fraction >= 0.50 and imp_data.shape[0] > 1) or (imp_data.shape[0] == 1)) and conc_gory != "<5 uM":  # the match ratio should be 50% or greater
                if len(set(list(uq_cps_dict2.values())[i][1])) == 1 and set(list(uq_cps_dict2.values())[i][1]) == {'-'}:
                    trend_str = "decrease"
                if len(set(list(uq_cps_dict2.values())[i][1])) == 1 and set(list(uq_cps_dict2.values())[i][1]) == {'+'}:
                    trend_str = "increase"
                if len(set(list(uq_cps_dict2.values())[i][1])) > 1:
                    trend_str = "different trend"
                if not '*' in list(uq_cps_dict2.values())[i][2]:  # the star numbers are all nan
                    star_str = "not significant"
                    starnumber_fraction = frac(0, 1)
                    starnumber_str = 'nan'
                if '*' in list(uq_cps_dict2.values())[i][2]:
                    star_str = "significant"
                    starnumber_fraction = frac(star_number, bottom)
                    starnumber_str = str(star_number) + '/' + str(bottom)
                if '!' in list(uq_cps_dict2.values())[i][3]:
                    signifi_str = 'FDR-signifi'

                ''' significant ratio select standard'''  #####

                if starnumber_fraction >= 0.5 and trend_str == 'increase':  # significant ratio must greater than 50%, to be considered as 'significant'
                    if signifi_str == 'FDR-signifi':
                        category = 'increase and FDR-signifi'
                        category_dict['increase and FDR-signifi'].append(cp_name)
                    else:
                        category = 'increase and significant'
                        category_dict['increase and significant'].append(cp_name)
                if starnumber_fraction >= 0.5 and trend_str == 'decrease':
                    if signifi_str == 'FDR-signifi':
                        category = 'decrease and FDR-signifi'
                        category_dict['decrease and FDR-signifi'].append(cp_name)
                    else:
                        category = 'decrease and significant'
                        category_dict['decrease and significant'].append(cp_name)
                if starnumber_fraction >= 0.5 and trend_str == 'different trend':
                    if signifi_str == 'FDR-signifi':
                        category = 'no trend and FDR-signifi'
                        category_dict['no trend and FDR-signifi'].append(cp_name)
                    else:
                        category = 'no trend and significant'
                        category_dict['no trend and significant'].append(cp_name)
                if starnumber_fraction < 0.5 and trend_str == 'decrease':
                    category_dict['decrease and not significant'].append(cp_name)
                    category = 'decrease and not significant'
                if starnumber_fraction < 0.5 and trend_str == 'increase':
                    category_dict['increase and not significant'].append(cp_name)
                    category = 'increase and not significant'
                if starnumber_fraction < 0.5 and trend_str == 'different trend':
                    category_dict['no trend and not significant'].append(cp_name)
                    category = 'no trend and not significant'

                row.append(fraction)  # match ratio (dec)
                row.append(str_fraction)  # match ratio (str)
                row.append(cp_name)  # cp name
                row.append(sorted(set(list(uq_cps_dict2.values())[i][0])))  # list ppms in data
                row.append(trend_str)  # in/decrease
                row.append(star_str)  # significance
                row.append(starnumber_str)  # significant ratio
                row.append(category)
                row.append(conc_gory)
                row = np.array(row, dtype=object)
                f_table2 = np.vstack((f_table2, row))
    f_table2 = np.delete(f_table2, 0, 0)
    metabolites_list = []
    metabolites_list = get_all_metabolites(category_dict, metabolites_list)
    f_table,_outputTable = format_f_table(f_table2)
    if args.param != 'None':
        #保存
        pd.DataFrame(_outputTable).to_excel(args.param,index=False)
    '''
    the disease related metabolites option
    '''
    if ans.strip().lower() == 'y':  # if they have a priority list
        if action.strip() == '1':
            filename = 'alzheimers_disease_metabolites.xlsx'
        elif action.strip() == '2':
            filename = 'lung_cancer_metabolites.xlsx'
        elif action.strip() == '3':
            filename = 'prostate_cancer_metabolites.xlsx'
        elif action.strip() == '0':
            filename = args.priorityfilename
        os.chdir(database_location)
        # direct to the disease list folder  # file location 5
        path = os.path.abspath(filename)
        df = pd.read_excel(path)
        priority_cp_list = list(sorted(set(np.asarray(df)[:, 0])))
        p_table = list(range(0, f_table2.shape[1]))
        output_dict3 = {}
        for i in range(0, len(priority_cp_list)):
            for j in range(0, len(output_dict2)):
                if priority_cp_list[i].strip().lower() == list(output_dict2.keys())[j].strip().lower():
                    output_dict3.update({list(output_dict2.keys())[j]: output_dict2[list(output_dict2.keys())[
                        j]]})  # add the metabolite in both output_dict2 and priority_cp_list to the output_dict3
        for i in range(0, len(priority_cp_list)):
            row = []
            for j in range(0, f_table2.shape[0]):
                if priority_cp_list[i].strip().lower() == f_table2[j, 2].strip().lower():
                    row.append(f_table2[j, 0])
                    row.append(f_table2[j, 1])
                    # add cp name
                    row.append(priority_cp_list[i])
                    row.append(f_table2[j, 3])
                    row.append(f_table2[j, 4])
                    row.append(f_table2[j, 5])
                    row.append(f_table2[j, 6])
                    row.append(f_table2[j, 7])
                    row.append(f_table2[j, 8])
            if len(row) == 0:  # if the cp doesn't appear in the input data
                row.append(0)
                row.append("")
                row.append("")
                # add cp name
                row.append(priority_cp_list[i])
                row.append("")
                row.append("")
                row.append("")
                row.append("")
                row.append("")
            row = np.asarray(row, dtype=object)
            p_table = np.vstack((p_table, row))
        # delete row [0] = ["","","",""]
        p_table = np.delete(p_table, 0, 0)
        p_table = sort_fractions(p_table)
        disease_metabolites_list = []
        for n in range(0, p_table.shape[0]):
            if len(p_table[n, 0]) != 0:
                disease_metabolites_list.append(p_table[n, 1])
        f_table2 = sort_fractions(f_table2)
        category_dict_p1 = turn_table_to_dict(p_table)
        category_dict_p2 = abbreviate_dict(category_dict_p1)
        category_dict_p2 = forma_dict(category_dict_p2)
        # os.chdir(output_location)
        '''
        output the result
        '''
        os.chdir(file_location)
        if category_type == '1':
            categorize_in_groups(category_dict_p2)
            forma_excel(filename_categorize)
        if category_type == '2':
            categorize_in_regions(category_dict_p2)
        if category_type == '3':
            output_the_list(disease_metabolites_list)
    else:
        # os.chdir(output_location)
        os.chdir(file_location)
        f_table2 = sort_fractions(f_table2)
        category_dict2 = abbreviate_dict(category_dict)
        category_dict2 = forma_dict(category_dict2)
        # os.chdir(output_location)
        if category_type == '1':
            categorize_in_groups(category_dict2)
            forma_excel(filename_categorize)
        if category_type == '2':
            categorize_in_regions(category_dict2)
        if category_type == '3':
            output_the_list(metabolites_list)


import argparse, os

# 当前目录
file_location = os.getcwd()
database_location = file_location + '\data base'
ppm_location = file_location + '\data base\metabolites_ppm_db'
region_location = file_location + '\data base\metabolites_region_db'
conc_location = file_location + '\data base\metabolites_conc_db'
print('\nMetabolites Search Program')

# 创建ArgumentParser对象
parser = argparse.ArgumentParser(description='计算调用')


# 添加参数
parser.add_argument('-database', type=str)
parser.add_argument('-coveredregion', type=str)
parser.add_argument('-filename', type=str)
parser.add_argument('-priorityfilename', default='', type=str)
parser.add_argument('-categorytype', type=str)
parser.add_argument('-ans', type=str)
parser.add_argument('-Action', default='', type=str)
parser.add_argument('-result', default='', type=str)
parser.add_argument('-abbreviation', default='None', type=str)
parser.add_argument('-param', default='None', type=str)


# 解析参数
args = parser.parse_args()

database = args.database
covered_region = args.coveredregion

# read peaks database, which is the ppm_table
os.chdir(ppm_location)
# directing to the database folder  # file location 1
filename = database + '_peaks_db.xlsx'
path = os.path.abspath(filename)
df = pd.read_excel(path)
ppm_table = np.asarray(df)
ppm_table0 = copy.deepcopy(ppm_table)

# read peakreagion database, which is the region_table
os.chdir(region_location)
# directing to the database folder  # file location 2
filename = database + '_region_db.xlsx'
path = os.path.abspath(filename)
df = pd.read_excel(path)
region_table = np.asarray(df)
region_table0 = copy.deepcopy(region_table)

# read concentration filter
os.chdir(conc_location)
filename = database + " conc list.xlsx"
path = os.path.abspath(filename)
df = pd.read_excel(path)
concfilter = np.asarray(df)

# read the abbreviation table
os.chdir(database_location)
# directing to the abbreviation folder # file location 3
filename = args.abbreviation
if filename == 'None':
    filename = 'abbreviation list.xlsx'
    path = os.path.abspath(filename)
    df = pd.read_excel(path)
    abbreviation_table = np.asarray(df)
else:
    os.chdir('../')
    path = os.getcwd() + "/" + filename
    df = pd.read_excel(path)
    abbreviation_table = np.asarray(df)
    if abbreviation_table.shape[1] >= 2:
        mask = abbreviation_table[:, 1] != 'nan'
        abbreviation_table = abbreviation_table[mask]  # remove the empty line
    else:
        abbreviation_table = np.array([["1","2"]])

# define which region is analysed
# covered_region = input(
#     "Please input the region to be analysed (If there are multiple regions please split with';', for example 8.5-8.0; 5.5-0.5): ")
# UI to import data (excel sheet)
os.chdir(file_location)
filename = args.filename
# directing to the Region of interes folder  # file location 4

path = os.path.abspath(filename)
df = pd.read_excel(path)
imp_data = np.asarray(df)

# print('format of the result:')
# print('(Press 1): put the metabolites into different test groups.')
# print('(Press 2): put the metabolites into different Regions of Interest')
# print('(Press 3): just output all the metabolites in a list')
category_type = args.categorytype

ans = args.ans
if ans.strip().lower() == 'y':  # if they have a priority list
    # print("(Press 1) Alzheimer's Disease Metabolites")
    # print("(Press 2) Lung Cancer Metabolites")
    # print("(Press 3) Prostate Cancer Metabolties ")
    # print("(Press 0) Enter own priority metabolites list")
    action = args.Action

# output_location = input('\nThe location of the output file: ')

filename_categorize = args.result
UI_search()

