import shutil
import sys,os,pandas as pd
# 导入图形组件库
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
#导入做好的界面库
from untitled import Ui_MainWindow
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class MainWindow(QMainWindow,Ui_MainWindow):
    def __init__(self):
        #Inherit the properties of the parent class (QMainWindow, Ui_MainWindow).
        super(MainWindow,self).__init__()
        #Initialize the interface components.
        self.setupUi(self)
        self.stackedWidget.setCurrentIndex(1)
        #initialization
        self.listWidget.addItems(
            [dirName for dirName in os.listdir("dataResult")]
        )

        #Introduction
        self.pushButton_4.clicked.connect(lambda :self.stackedWidget.setCurrentIndex(1))
        #Add examples
        self.pushButton.clicked.connect(self.addList)
        self.pushButton_2.clicked.connect(self.deleteList)

        #Add line
        self.pushButton_7.clicked.connect(self.addRow)
        #Delete line
        self.pushButton_8.clicked.connect(self.removeRow)

        self.comboBox.addItems([
            'csf','feces','saliva','serum','sweat','urine'
        ])

        #yes or no
        self.buttonGroup.buttonToggled.connect(self.changeState)
        #choose excel
        self.pushButton_10.clicked.connect(self.loadExcel)

        #calculate
        self.pushButton_11.clicked.connect(self.calValue)

        #double click
        self.listWidget.doubleClicked.connect(self.showAna)

        #combine
        self.pushButton_3.clicked.connect(self.combineResult)

        self.listWidget.setSelectionMode(QListWidget.MultiSelection)
        self.pushButton_6.clicked.connect(self.close)

        #calculate
        self.pushButton_9.clicked.connect(self.calFinal)

        #export
        self.pushButton_5.clicked.connect(self.outputExcel)

        pixmap = QPixmap('Introduction.jpg')
        target_width = 950
        original_ratio = pixmap.width() / pixmap.height()
        new_height = int(target_width / original_ratio)
        resize_pixmap = pixmap.scaled(target_width, new_height, Qt.KeepAspectRatio)
        self.label_11.setPixmap(QPixmap(resize_pixmap.toImage()))

        # Connect the itemClicked signal to the slot function.
        self.listWidget.itemClicked.connect(self.onItemClicked)


    def onItemClicked(self, item):
        selected_items = self.listWidget.selectedItems()
        if len(selected_items) > 2:
            # If more than two options are selected, reset the first two selections.
            for selected_item in selected_items[:2]:
                selected_item.setSelected(False)
    def outputExcel(self):
        wb = Workbook()
        ws = wb.active

        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                if item:
                    cell = ws.cell(row=row + 1, column=col + 1, value=item.text())
                    if item.font().overline():
                        # set the font
                        font = Font(name=item.font().family(),
                                    size=item.font().pointSize(),
                                    bold=item.font().bold(),
                                    italic=item.font().italic(),
                                    underline='double',  # double underline
                                    color=self.convert_to_argb(item.foreground().color()))
                    else:
                        # set the font
                        font = Font(name=item.font().family(),
                                    size=item.font().pointSize(),
                                    bold=item.font().bold(),
                                    italic=item.font().italic(),
                                    underline='single' if item.font().underline() else None,  # underline
                                    color=self.convert_to_argb(item.foreground().color()))
                    cell.font = font
                    # set background color
                    if self.convert_to_argb(item.background().color()) == '00000000':
                        fill = PatternFill(start_color='ffffffff',
                                           end_color='ffffffff',
                                           fill_type='solid')
                    else:
                        fill = PatternFill(start_color=self.convert_to_argb(item.background().color()),
                                           end_color=self.convert_to_argb(item.background().color()),
                                           fill_type='solid')
                    cell.fill = fill

                    # Set alignment mode.
                    alignment = Alignment(horizontal='center', vertical='center')
                    cell.alignment = alignment

        # Adjust the column width.
        for col in range(self.tableWidget.columnCount()):
            column_letter = get_column_letter(col + 1)
            ws.column_dimensions[column_letter].width = 15


        fileName_choose, filetype = QFileDialog.getSaveFileName(self,
                                                                "Output To",
                                                                "./",  # path
                                                                "Excel file (*.xlsx)")
        if fileName_choose:
            wb.save(fileName_choose)
            QMessageBox.information(self,'Tips','Success',QMessageBox.Yes)

    def convert_to_argb(self, qcolor):
        """
        Convert QColor to an ARGB-formatted hexadecimal string.
        aRGB format：AARRGGBB（Alpha, Red, Green, Blue）
        """
        alpha = qcolor.alpha()  # acquire Alpha tunnel
        red = qcolor.red()
        green = qcolor.green()
        blue = qcolor.blue()

        if red == 0 and green == 0 and blue == 0:
            return f"{0:02X}{0:02X}{0:02X}{0:02X}"

        return f"{alpha:02X}{red:02X}{green:02X}{blue:02X}"

    def calFinal(self):
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        data5 = []
        # 读取数据
        for col in range(self.tableWidget_2.columnCount()):
            for row in range(1, self.tableWidget_2.rowCount()):
                if col == 0:
                    value = self.tableWidget_2.item(row, col).text()
                    data1.append(value)
                elif col == 1:
                    value = self.tableWidget_2.item(row, col).text()
                    data2.append(value)
                elif col == 2:
                    value = self.tableWidget_2.item(row, col).text()
                    data3.append(value)
                elif col == 3:
                    value = self.tableWidget_2.item(row, col).text()
                    data4.append(value)
                elif col == 4:
                    value = self.tableWidget_2.item(row, col).text()
                    data5.append(value)
        _filename = os.path.join(self.path, self._name + ".xlsx")
        pd.DataFrame({
            'ppm upp lim': data1,
            'ppm low lim': data2,
            'decrease/increase': data3,
            'signficiance': data4,
            'FDR': data5
        }).to_excel(_filename, index=False)

        _reusltPath1 = os.path.join(self.path, "reult1.xlsx")
        _reusltPath2 = os.path.join(self.path, "reult2.xlsx")

        _reusltPath3 = os.path.join(self.path, "output.xlsx")



        #calculate
        _outputTable = {
            "metabolites": [],
            "abbreviations": [],
            "match ratio": [],
            'matched regions': [],
            "concentration range": []
        }
        for row in range(self.tableWidget_4.rowCount()):
            _outputTable['metabolites'].append(self.tableWidget_4.item(row, 0).text())
            _outputTable['abbreviations'].append(self.tableWidget_4.item(row, 1).text())
            _outputTable['match ratio'].append(self.tableWidget_4.item(row, 2).text())
            _outputTable['matched regions'].append(self.tableWidget_4.item(row, 3).text())
            _outputTable['concentration range'].append(self.tableWidget_4.item(row, 4).text())

        pd.DataFrame(_outputTable).to_excel(_reusltPath3,index=False)

        if self.radioButton_2.isChecked():
            # calculate
            os.system(
                f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans n -result {_reusltPath1} -abbreviation {_reusltPath3}')
            os.system(
                f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 2 -ans n -result {_reusltPath2} -abbreviation {_reusltPath3}')
        else:
            if self.radioButton_3.isChecked():
                # calculate
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 1 -result {_reusltPath1} -abbreviation {_reusltPath3}')
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 2 -ans y -Action 1 -result {_reusltPath2} -abbreviation {_reusltPath3}')
            elif self.radioButton_4.isChecked():
                # calculate
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 2 -result {_reusltPath1} -abbreviation {_reusltPath3}')
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 2 -ans y -Action 2 -result {_reusltPath2} -abbreviation {_reusltPath3}')
            elif self.radioButton_5.isChecked():
                # calculate
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 3 -result {_reusltPath1} -abbreviation {_reusltPath3}')
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 2 -ans y -Action 3 -result {_reusltPath2} -abbreviation {_reusltPath3}')
            else:
                # calculate
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 0 -priorityfilename {self.lineEdit_3.text()} -result {_reusltPath1} -abbreviation {_reusltPath3}')
                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 2 -ans y -Action 0 -priorityfilename {self.lineEdit_3.text()} -result {_reusltPath2} -abbreviation {_reusltPath3}')


        groups = [self._name]
        datas = [self.loadExcelSheet(_reusltPath1)]
        self.showResult(self.tableWidget, datas, groups)

        self.showTable1(self.tableWidget_3, _reusltPath2)
        # Save results.
        with open(os.path.join(self.path, "param.txt"), 'w') as f:
            f.write(str(self.comboBox.currentIndex()) + "\n")
            f.write(f'{self.lineEdit.text()}-{self.lineEdit_2.text()}\n')
            if self.radioButton_2.isChecked():
                f.write('no')
            else:
                f.write('yes\n')
                if self.radioButton_3.isChecked():
                    f.write('1')
                elif self.radioButton_4.isChecked():
                    f.write('2')
                elif self.radioButton_5.isChecked():
                    f.write('3')
                elif self.radioButton_7.isChecked():
                    f.write('4\n')
                    f.write(f'{self.lineEdit_3.text()}')

        QMessageBox.information(self, "Tips", "Calculate successfully", QMessageBox.Yes)

    def combineResult(self):
        self.stackedWidget.setCurrentIndex(0)
        self.tabWidget.setCurrentIndex(1)

        # get the selected element
        selected_items = self.listWidget.selectedItems()

        # Retrieve the names of the selected items and display them.
        selected_names = [item.text() for item in selected_items]

        # Display the selected item(s).
        if selected_names:
            groups = [item.text() for item in selected_items]
            datas = [self.loadExcelSheet(os.path.join(os.path.join('dataResult',groups[0]), "reult1.xlsx")),self.loadExcelSheet(os.path.join(os.path.join('dataResult',groups[1]), "reult1.xlsx"))]
            self.showResult(self.tableWidget, datas, groups)
            self.tableWidget_3.setVisible(False)
            self.pushButton_5.setVisible(True)
            self.tabWidget.setTabVisible(0,False)
            self.tabWidget.setTabVisible(1,False)
        else:
            QMessageBox.information(self, "No Selection", "No items selected.")


    def showAna(self):
        self.tabWidget.setTabVisible(0, True)
        self.tabWidget.setTabVisible(1, True)
        self.stackedWidget.setCurrentIndex(0)
        self.tabWidget.setCurrentIndex(0)


        self._name = self.listWidget.currentItem().text()
        self.path = os.path.join('dataResult',self._name)

        _reusltPath3 = os.path.join(self.path, "output.xlsx")
        try:
            self.showTable1(self.tableWidget_4, _reusltPath3)
        except:
            pass

        try:
            try:
                _df = pd.read_excel(os.path.join(self.path,self._name + ".xlsx"))
                self.showTable(self.tableWidget_2, _df.values.tolist())
            except:
                self.showTable(self.tableWidget_2, [])
            #display
            with open(os.path.join(self.path, "param.txt"), 'r') as f:
                _params = f.read().split('\n')

            self.comboBox.setCurrentIndex(int(_params[0]))
            self.lineEdit.setText(_params[1].split('-')[0])
            self.lineEdit_2.setText(_params[1].split('-')[1])

            if _params[2] == 'no':
                self.radioButton_2.setChecked(True)
            else:
                if _params[3] == '1':
                    self.radioButton_3.setChecked(True)
                elif _params[3] == '2':
                    self.radioButton_4.setChecked(True)
                elif _params[3] == '3':
                    self.radioButton_5.setChecked(True)
                elif _params[3] == '4':
                    self.radioButton_7.setChecked(True)
                    self.lineEdit_3.setText(_params[4])

            _reusltPath1 = os.path.join(self.path, "reult1.xlsx")

            if os.path.exists(_reusltPath1):
                groups = [self._name]
                datas = [self.loadExcelSheet(_reusltPath1)]
                self.showResult(self.tableWidget, datas, groups)
                try:
                    _reusltPath2 = os.path.join(self.path, "reult2.xlsx")
                    self.showTable1(self.tableWidget_3, _reusltPath2)
                except:
                    pass
            self.tableWidget_3.setVisible(True)
            self.pushButton_5.setVisible(False)
        except Exception as e:
            # initialize the right part
            self.tableWidget_2.clear()
            self.tableWidget.clear()
            self.tableWidget.setRowCount(0)  # Set the number of rows in the table.
            self.tableWidget_2.setRowCount(0)  #
            self.comboBox.setCurrentIndex(0)
            self.lineEdit_3.clear()
            self.lineEdit.clear()
            self.lineEdit_2.clear()
            self.radioButton_2.setChecked(True)
            self.tableWidget_3.setVisible(True)
            self.pushButton_5.setVisible(False)

            self.showTable(self.tableWidget_2, [])
            self.tableWidget.clear()
            self.tableWidget_3.clear()
            self.tableWidget_4.clear()

    def deleteList(self):
        _name = self.listWidget.currentItem().text()
        path = os.path.join('dataResult', _name)
        shutil.rmtree(path)
        self.listWidget.clear()
        self.listWidget.addItems(
            [dirName for dirName in os.listdir("dataResult")]
        )
        self.stackedWidget.setCurrentIndex(1)

    def loadExcelSheet(self,path):
        datas = []
        # read Excel file
        wb = load_workbook(path)
        sheet = wb.active

        num = 0
        # Iterate through all cells.
        for row in sheet.iter_rows():
            num += 1
            if num != 1:
                data = []
                for cell in row[1:]:
                    # the content of cell
                    cell_value = cell.value

                    # the background color of cell
                    background_color = cell.fill.start_color.rgb

                    # font color
                    font_color = cell.font.color
                    #Check if font_color is None and whether font_color.rgb is valid.
                    if font_color and font_color.rgb:
                        font_rgb = font_color.rgb
                    else:
                        font_rgb = None

                    # check whether font is italic
                    italic = cell.font.italic

                    data.append(
                        [cell_value,background_color,font_rgb,italic]
                    )
                datas.append(data)

        return datas



    def showResult(self, tableWidget, datas, groups):

        col_name = [
            'Groups',
            'Decrease',
            'Increase'
        ]
        tableWidget.clear()
        tableWidget.setSelectionMode(QTableWidget.NoSelection)

        _nums = 0
        for data in datas:
            _nums += len(data)

        tableWidget.setRowCount(_nums + 1)  # set the row number of table
        tableWidget.setColumnCount(9)  # set the column number of table
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Set horizontal and vertical labels to be hidden.
        tableWidget.verticalHeader().setVisible(False)
        tableWidget.horizontalHeader().setVisible(False)

        #The first column is not set
        item = QTableWidgetItem(col_name[0])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 0, item)


        #Columns two to five
        tableWidget.setSpan(0, 1, 1, 4)
        item = QTableWidgetItem(col_name[1])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 1, item)


        # columns 6 to nine
        tableWidget.setSpan(0, 5, 1, 4)
        item = QTableWidgetItem(col_name[2])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 5, item)

        for i,group in enumerate(groups):
            if i == 0:
                tableWidget.setSpan(1, 0, len(datas[i]), 1)
                item = QTableWidgetItem(group)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                tableWidget.setItem(i+1, 0, item)
            else:
                tableWidget.setSpan(len(datas[0]) + 1, 0, len(datas[i]), 1)
                item = QTableWidgetItem(group)
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                tableWidget.setItem(len(datas[0]) + 1, 0, item)

        def find_duplicates_list_comprehension(lst):
            return list({x for x in lst if lst.count(x) > 1 and x and x.strip() !=''} )

        #data loading
        if len(datas) == 1:
            #all data
            dataInFirst = datas[0]
            # whether it's underlined
            _data1 = []
            _data2 = []
            for i in range(len(dataInFirst)):
                for m in range(4):
                    _data1.append(dataInFirst[i][m][0])

            for i in range(len(dataInFirst)):
                for m in range(4,8):
                    _data2.append(dataInFirst[i][m][0])
            c1 = find_duplicates_list_comprehension(_data1)
            c2 = find_duplicates_list_comprehension(_data2)
            common = list(set(_data1) & set(_data2))
            common = [i for i in common if i and i.strip() != '']
            for i in range(len(dataInFirst)):
                for m in range(8):
                    _contents = dataInFirst[i][m]
                    if _contents[0]:
                        item = QTableWidgetItem(str(_contents[0]))
                    else:
                        item = QTableWidgetItem('')
                    item.setBackground(QColor("#"+str(_contents[1])[2:]))
                    item.setForeground(QColor("#"+str(_contents[2])[2:]))

                    font = QFont()
                    if _contents[3]:
                        font.setItalic(True)
                    if _contents[0]:
                        if _contents[0] in common:
                            font.setUnderline(True)

                    if _contents[0]:
                        if _contents[0] in c1 or _contents[0] in c2:
                            font.setOverline(True)  # overline
                    item.setFont(font)

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    tableWidget.setItem(i + 1, m+1, item)  # row and columns




        else:
            # all data
            dataInFirst = datas[0]
            _dataFirstLastColor = dataInFirst[-1][0][1]
            _white1 = []
            for i in range(len(dataInFirst)):
                for m in range(8):
                    if dataInFirst[i][m][1] == _dataFirstLastColor:
                        _white1.append(dataInFirst[i][m][0])

            dataInSecond = datas[1]
            _dataSecondLastColor = dataInSecond[-1][0][1]
            _white2 = []
            for i in range(len(dataInSecond)):
                for m in range(8):
                    if dataInSecond[i][m][1] == _dataSecondLastColor:
                        _white2.append(dataInSecond[i][m][0])
            # print(_white1)
            # print(_white2)
            # print(len(dataInFirst),len(dataInSecond))
            if len(dataInFirst) >= len(dataInSecond):

                # whether underlining
                _data1 = []
                _data2 = []
                for i in range(len(dataInFirst)):
                    for m in range(4):

                        try:
                            _data1.append(dataInFirst[i][m][0])
                        except:
                            pass
                        try:
                            _data1.append(dataInSecond[i][m][0])
                        except:
                            pass

                for i in range(len(dataInFirst)):
                    for m in range(4, 8):
                        try:
                            _data2.append(dataInFirst[i][m][0])
                        except:
                            pass
                        try:
                            _data2.append(dataInSecond[i][m][0])
                        except:
                            pass
            else:
                # whether underlining
                _data1 = []
                _data2 = []
                for i in range(len(dataInSecond)):

                    for m in range(4):
                        try:
                            _data1.append(dataInFirst[i][m][0])
                        except:
                            pass
                        try:
                            _data1.append(dataInSecond[i][m][0])
                        except:
                            pass

                for i in range(len(dataInSecond)):
                    for m in range(4, 8):
                        try:
                            _data2.append(dataInFirst[i][m][0])
                        except:
                            pass
                        try:
                            _data2.append(dataInSecond[i][m][0])
                        except:
                            pass

            common = list(set(_data1) & set(_data2))
            commons = [i for i in common if i and i.strip() != '']


            c11 = find_duplicates_list_comprehension(_data1)
            c22 = find_duplicates_list_comprehension(_data2)

            common = []
            for m in commons:
                if m in _white1 and m in _white2:
                    pass
                else:
                    common.append(m)

            c1 = []
            for m in c11:
                if m in _white1 and m in _white2:
                    pass
                else:
                    c1.append(m)

            c2 = []
            for m in c22:
                if m in _white1 and m in _white2:
                    pass
                else:
                    c2.append(m)

            for i in range(len(dataInFirst)):
                for m in range(8):
                    _contents = dataInFirst[i][m]
                    if _contents[0]:
                        item = QTableWidgetItem(str(_contents[0]))
                    else:
                        item = QTableWidgetItem('')
                    item.setBackground(QColor("#" + str(_contents[1])[2:]))
                    item.setForeground(QColor("#" + str(_contents[2])[2:]))

                    font = QFont()
                    if _contents[3]:
                        font.setItalic(True)
                    if _contents[0]:
                        if _contents[0] in common:
                            font.setUnderline(True)  # underline
                    #
                    if _contents[0]:
                        if _contents[0] in c1 or _contents[0] in c2:
                            font.setOverline(True)  # overline


                    item.setFont(font)

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    tableWidget.setItem(i + 1, m + 1, item)



            for i in range(len(dataInSecond)):
                for m in range(8):
                    _contents = dataInSecond[i][m]
                    if _contents[0]:
                        item = QTableWidgetItem(str(_contents[0]))
                    else:
                        item = QTableWidgetItem('')
                    item.setBackground(QColor("#" + str(_contents[1])[2:]))
                    item.setForeground(QColor("#" + str(_contents[2])[2:]))

                    font = QFont()
                    if _contents[3]:
                        font.setItalic(True)
                    if _contents[0]:
                        if _contents[0] in common:
                            font.setUnderline(True)

                    if _contents[0]:
                        if _contents[0] in c1 or _contents[0] in c2:
                            font.setOverline(True)

                    item.setFont(font)

                    item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    tableWidget.setItem(len(dataInFirst)+i + 1, m + 1, item)



    def calValue(self):
        data1 = []
        data2 = []
        data3 = []
        data4 = []
        data5 = []
        #read data
        for col in range(self.tableWidget_2.columnCount()):
            for row in range(1,self.tableWidget_2.rowCount()):
                if col == 0:
                    value = self.tableWidget_2.item(row, col).text()
                    data1.append(value)
                elif col == 1:
                    value = self.tableWidget_2.item(row, col).text()
                    data2.append(value)
                elif col == 2:
                    value = self.tableWidget_2.item(row, col).text()
                    data3.append(value)
                elif col == 3:
                    value = self.tableWidget_2.item(row, col).text()
                    data4.append(value)
                elif col == 4:
                    value = self.tableWidget_2.item(row, col).text()
                    data5.append(value)

        _filename = os.path.join(self.path,self._name + ".xlsx")
        pd.DataFrame({
            'ppm upp lim':data1,
            'ppm low lim':data2,
            'decrease/increase':data3,
            'signficiance':data4,
            'FDR':data5
        }).to_excel(_filename,index=False)

        _reusltPath1 = os.path.join(self.path, "reult1.xlsx")

        _reusltPath3 = os.path.join(self.path, "output.xlsx")

        if self.radioButton_2.isChecked():

            os.system(f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans n -result {_reusltPath1} -param {_reusltPath3}')

        else:
            if self.radioButton_3.isChecked():

                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 1 -result {_reusltPath1} -param {_reusltPath3}')

            elif self.radioButton_4.isChecked():

                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 2 -result {_reusltPath1} -param {_reusltPath3}')

            elif self.radioButton_5.isChecked():

                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 3 -result {_reusltPath1} -param {_reusltPath3}')

            else:

                os.system(
                    f'python func.py -database {self.comboBox.currentText()} -coveredregion {self.lineEdit.text()}-{self.lineEdit_2.text()} -filename {_filename} -categorytype 1 -ans y -Action 0 -priorityfilename {self.lineEdit_3.text()} -result {_reusltPath1} -param {_reusltPath3}')

        self.showTable1(self.tableWidget_4,_reusltPath3)
        # save result
        with open(os.path.join(self.path, "param.txt"), 'w') as f:
            f.write(str(self.comboBox.currentIndex()) + "\n")
            f.write(f'{self.lineEdit.text()}-{self.lineEdit_2.text()}\n')
            if self.radioButton_2.isChecked():
                f.write('no')
            else:
                f.write('yes\n')
                if self.radioButton_3.isChecked():
                    f.write('1')
                elif self.radioButton_4.isChecked():
                    f.write('2')
                elif self.radioButton_5.isChecked():
                    f.write('3')
                elif self.radioButton_7.isChecked():
                    f.write('4\n')
                    f.write(f'{self.lineEdit_3.text()}')

        QMessageBox.information(self,"Tips","Calculate successfully",QMessageBox.Yes)

    def showTable1(self, tableWidget, path):
        df = pd.read_excel(path)
        col_name, data  = df.columns.tolist(),df.values.tolist()
        tableWidget.clear()
        tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)

        header_font = QFont("微软雅黑", 12, QFont.Bold)  # Using Microsoft Yahei font, set the size and thickness
        tableWidget.horizontalHeader().setFont(header_font)  # Sets the header
        tableWidget.setRowCount(len(data))  # Sets the number of rows for the table
        tableWidget.setColumnCount(len(col_name))  # Sets the number of columns for the table
        tableWidget.setHorizontalHeaderLabels(col_name)
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tableWidget.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: black;       /* 黑色背景 */
                color: white;                  /* 白色字体 */
                border-bottom: 2px solid black; /* 黑色下划线 */
            }
        """)

        tableWidget.verticalHeader().setVisible(False)
        for i in range(len(data)):
            for m in range(len(col_name)):
                if str(data[i][m]) == 'nan':
                    newItem = QTableWidgetItem('')
                else:
                    newItem = QTableWidgetItem(str(data[i][m]))

                newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                tableWidget.setItem(i, m, newItem)

    def changeState(self):
        if self.radioButton_2.isChecked():
            self.widget_2.setVisible(False)
        else:
            self.widget_2.setVisible(True)


    def loadExcel(self):
        fileName1, filetype = QFileDialog.getOpenFileName(self,
                                                          "选取Excel文件",
                                                          "./",
                                                          "Excel File (*.xlsx;*.csv)")  # Set file extension filtering, using double semicolons
        if fileName1:
            self.lineEdit_3.setText(fileName1)


    def addRow(self):
        """增加一行"""
        current_row_count = self.tableWidget_2.rowCount()  # current row number
        self.tableWidget_2.insertRow(current_row_count)  # insert one row in the end
        self.centerAllCells()

    def removeRow(self):
        """删除一行"""
        current_row_count = self.tableWidget_2.selectedItems()  # current row
        _rows = []
        for item in current_row_count:
            _rows.append(item.row())
        _rows = list(set(_rows))
        if len(_rows) == 0:
            self.tableWidget_2.removeRow(self.tableWidget_2.rowCount() - 1)
        else:
            for i in sorted(_rows,reverse=True):
                self.tableWidget_2.removeRow(i)

        self.centerAllCells()

    def centerAllCells(self):
        """使所有单元格内容居中"""
        row_count = self.tableWidget_2.rowCount()
        column_count = self.tableWidget_2.columnCount()
        for row in range(row_count):
            for col in range(column_count):
                item = self.tableWidget_2.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignCenter)
                else:
                    newItem = QTableWidgetItem('')
                    newItem.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
                    self.tableWidget_2.setItem(row,col,newItem)

    def showTable(self, tableWidget, data):
        col_name = [
            'Regions of interest(ppm)',
            'decrease/increase',
            'Significance 1',
            'Significance 2'
        ]
        tableWidget.clear()
        tableWidget.setSelectionBehavior(QAbstractItemView.SelectRows)

        tableWidget.setRowCount(len(data) + 1)
        tableWidget.setColumnCount(5)  #
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tableWidget.verticalHeader().setVisible(False)
        tableWidget.horizontalHeader().setVisible(False)
        tableWidget.setSpan(0, 0, 1, 2)

        item = QTableWidgetItem(col_name[0])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 0, item)

        item = QTableWidgetItem(col_name[1])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 2, item)

        item = QTableWidgetItem(col_name[2])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 3, item)

        item = QTableWidgetItem(col_name[3])
        item.setTextAlignment(Qt.AlignCenter)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        tableWidget.setItem(0, 4, item)

        for i in range(len(data)):
            for m in range(len(col_name)+1):

                try:
                    if str(data[i][m]) == 'nan':
                        newItem = QTableWidgetItem('')
                    else:
                        newItem = QTableWidgetItem(str(data[i][m]))
                except:
                    newItem = QTableWidgetItem('')
                tableWidget.setItem(i+1, m, newItem)
        self.centerAllCells()

    def keyPressEvent(self, event):
        """capture the action of Ctrl+V"""
        if event.matches(QKeySequence.Paste):  # Check whether it is Ctrl+V
            self.pasteExcelData()

    def pasteExcelData(self):
        # Get clipboard content
        clipboard = QApplication.clipboard()
        data = clipboard.text()

        # Split the content by row and column
        rows = data.strip().split('\n')

        datas = []
        for row_index, row in enumerate(rows):
            columns = row.split('\t')
            datas.append(columns)

        self.showTable(self.tableWidget_2,datas)

    def addList(self):
        text, okPressed = QInputDialog.getText(self, "New", "name:", QLineEdit.Normal, "")
        if okPressed and text != '':
            _path = os.path.join('dataResult', text)
            if os.path.exists(_path):
                QMessageBox.warning(self, "Warning", "Exist this analysis", QMessageBox.Yes)
            else:
                self.listWidget.addItem(text)
                self.stackedWidget.setCurrentIndex(0)

                self.tabWidget.setCurrentIndex(0)

                self.tableWidget_2.clear()
                self.tableWidget.clear()
                self.tableWidget.setRowCount(0)
                self.tableWidget_2.setRowCount(0)
                self.comboBox.setCurrentIndex(0)
                self.lineEdit_3.clear()
                self.lineEdit.clear()
                self.lineEdit_2.clear()
                self.radioButton_2.setChecked(True)
                self.tabWidget.setTabVisible(0, True)
                self.tabWidget.setTabVisible(1, True)
                self.tableWidget_3.setVisible(True)
                self.pushButton_5.setVisible(False)

                os.makedirs(_path)
                self.path = _path
                self._name = text

                self.showTable(self.tableWidget_2,[])

                self.tableWidget.clear()
                self.tableWidget_3.clear()
                self.tableWidget_4.clear()


if __name__ == "__main__":
    QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling)
    #Create QApplication (fixed/common syntax)
    app = QApplication(sys.argv)
    # Instantiate the interface.
    window = MainWindow()
    #display interface
    window.show()
    sys.exit(app.exec_())